#!/user/bin/env python3
# -*- coding: utf-8 -*-
# @Author : Tikyo
# @Time   : 2019/6/26 19:11
from openpyxl import load_workbook
from handle_config import config
from handle_log import my_logger
from constance import TIME_RECORD_PATH
import os


class HandleExcel:
    head = ['CycleNo', 'LoadCosumablesTime', 'TransferTime', 'PushCosumablesTime']

    def __init__(self, file_name, sheet_name=None):
        self.file_name, self.sheet_name = file_name, sheet_name
        self.sheet_head = None

        # # 检查下文件是否存在，若不存在，则创建一个
        # if os.path.exists(TIME_RECORD_PATH):
        #     pass
        # else:
        #     os.mknod(TIME_RECORD_PATH)
        # my_wb = load_workbook(self.file_name)
        # if self.sheet_name is None:
        #     my_ws = my_wb.active
        # else:
        #     my_ws = my_wb[self.sheet_name]
        # col = 1
        # for data in self.head:
        #     my_ws.cell(row=1, column=col, value=data)
        #     col += 1
        # my_wb.save(self.file_name)
        # my_wb.close()

    def get_cases(self):
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]

        sheet_head_tuple = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        self.sheet_head = sheet_head_tuple[0]

        sheet_content = tuple(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        case_list = []
        for data in sheet_content:
            dict1 = dict(zip(self.sheet_head, data))
            case_list.append(dict1)
        return case_list

    def write_result(self, row, actual, result):
        # 可以跟get_cases共用wb吗？--不可，前面写入的数据可能没有保存
        my_wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            my_ws = my_wb.active
        else:
            my_ws = my_wb[self.sheet_name]
        if isinstance(row, int) and (2 <= row <= my_ws.max_row):
            my_ws.cell(row=row, column=config.get_int('excel', 'actual_col'), value=actual)
            my_ws.cell(row=row, column=config.get_int('excel', 'result_col'), value=result)
            my_wb.save(self.file_name)
            my_wb.close()
        else:
            print('传入行号有误')

    def record_time(self, row, col, result):
        my_wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            my_ws = my_wb.active
        else:
            my_ws = my_wb[self.sheet_name]
        if isinstance(row, int) and row > 1:
            my_ws.cell(row=row, column=col, value=result)
            my_wb.save(self.file_name)
            my_wb.close()
        else:
            my_logger.error('传入行号有误')

    def get_max_col(self):
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        return ws.max_column

    def get_cell_value(self, row, col):
        wb = load_workbook(self.file_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        return ws.cell(row=row, column=col).value


if __name__ == '__main__':
    haha = HandleExcel(TIME_RECORD_PATH)
    haha.record_time(2, 1, '00:00:20')
